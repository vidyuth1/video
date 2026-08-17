"""
Mold Coordinate Tracker
=======================
Streamlit app that lets you upload up to 40 photos of molds (or plates) laid
out as a 15 x 8 grid (120 coordinates), auto-prompts calibration on the first
image, then steps through each image with Next/Previous buttons.

Coordinate naming: columns A-O (15), rows 1-8 (8), e.g. "B3" = column B, row 3.

How it works
------------
1. Upload up to MAX_IMAGES images at once.
2. The app automatically opens the first image and prompts calibration.
3. Click the OUTER TOP-LEFT corner of the grid (just outside A1), then the
   OUTER BOTTOM-RIGHT corner (just outside O8). That rectangle is divided
   evenly into 15 × 8 cells.
4. After calibration the mode switches to "Mark coordinates". Click ANYWHERE
   inside a cell to toggle it between Present (green) and Empty (red). Every
   click is written to disk immediately.
5. Click "Next" to move to the next image (or "Previous" to go back).
6. After all images are reviewed, use the Export tab to download:
   - A heatmap image showing missing-coordinate frequency across all molds.
   - An Excel workbook with two sheets: the heatmap data and a frequency table.
"""

import hashlib
import io
import json
import os
from io import BytesIO

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import numpy as np
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
from PIL import Image, ImageDraw, ImageFont
from streamlit_image_coordinates import streamlit_image_coordinates

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
COLS = 15
ROWS = 8
COL_LABELS = "ABCDEFGHIJKLMNO"
MAX_DISPLAY_WIDTH = 900
MAX_IMAGES = 40

STATE_DIR = "well_data"
os.makedirs(STATE_DIR, exist_ok=True)

PRESENT_FILL = (46, 204, 113)
EMPTY_FILL   = (231, 76,  60)
CALIB_COLOR  = (52,  152, 219)
GRID_LINE    = (0,   0,   0, 180)

st.set_page_config(page_title="Mold Coordinate Tracker", layout="wide")

# ---------------------------------------------------------------------------
# PROFESSIONAL STYLING — injected once at startup
# ---------------------------------------------------------------------------
st.markdown("""
<style>
/* ── Google Font import ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

/* ── Global typography ── */
html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    letter-spacing: -0.01em;
}

/* ── Page background ── */
.stApp {
    background-color: #F7F8FA;
}

/* ── Main content area ── */
.main .block-container {
    padding-top: 2.5rem;
    padding-bottom: 3rem;
    max-width: 1100px;
}

/* ── Page title ── */
h1 {
    font-size: 1.65rem !important;
    font-weight: 700 !important;
    color: #111827 !important;
    letter-spacing: -0.03em !important;
    margin-bottom: 0.15rem !important;
}

/* ── Section headings ── */
h2, h3 {
    font-weight: 600 !important;
    color: #1F2937 !important;
    letter-spacing: -0.02em !important;
}

/* ── Caption / meta text ── */
.stCaption, small {
    color: #6B7280 !important;
    font-size: 0.82rem !important;
}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background-color: #FFFFFF;
    border-right: 1px solid #E5E7EB;
}

section[data-testid="stSidebar"] .block-container {
    padding-top: 2rem;
}

section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    font-size: 0.85rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
    color: #9CA3AF !important;
}

/* ── Buttons ── */
.stButton > button {
    border-radius: 6px !important;
    font-weight: 500 !important;
    font-size: 0.84rem !important;
    letter-spacing: 0em !important;
    border: 1px solid #D1D5DB !important;
    background-color: #FFFFFF !important;
    color: #374151 !important;
    transition: background 0.15s, border-color 0.15s, box-shadow 0.15s !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
    padding: 0.35rem 0.9rem !important;
}

.stButton > button:hover {
    background-color: #F3F4F6 !important;
    border-color: #9CA3AF !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08) !important;
}

.stButton > button[kind="primary"] {
    background-color: #1D4ED8 !important;
    color: #FFFFFF !important;
    border-color: #1D4ED8 !important;
}

.stButton > button[kind="primary"]:hover {
    background-color: #1E40AF !important;
    border-color: #1E40AF !important;
}

/* ── Download buttons ── */
.stDownloadButton > button {
    border-radius: 6px !important;
    font-weight: 500 !important;
    font-size: 0.84rem !important;
    border: 1px solid #D1D5DB !important;
    background-color: #FFFFFF !important;
    color: #374151 !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
}

.stDownloadButton > button:hover {
    background-color: #F9FAFB !important;
    border-color: #9CA3AF !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    border-bottom: 1px solid #E5E7EB;
    gap: 0;
}

.stTabs [data-baseweb="tab"] {
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    color: #6B7280 !important;
    padding: 0.65rem 1.1rem !important;
    border-radius: 0 !important;
    border-bottom: 2px solid transparent !important;
    background: transparent !important;
    letter-spacing: 0 !important;
}

.stTabs [aria-selected="true"] {
    color: #1D4ED8 !important;
    border-bottom-color: #1D4ED8 !important;
    font-weight: 600 !important;
}

/* ── Info / warning / success banners ── */
.stAlert {
    border-radius: 6px !important;
    font-size: 0.875rem !important;
    border-left-width: 3px !important;
}

/* ── Metrics ── */
[data-testid="metric-container"] {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 8px;
    padding: 0.75rem 1rem;
    box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}

[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-size: 1.5rem !important;
    font-weight: 700 !important;
    color: #111827 !important;
}

[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
    color: #6B7280 !important;
}

/* ── Progress bar ── */
.stProgress > div > div {
    border-radius: 99px !important;
    height: 3px !important;
    background-color: #E5E7EB !important;
}

.stProgress > div > div > div {
    background-color: #1D4ED8 !important;
    border-radius: 99px !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    border-radius: 8px !important;
}

[data-testid="stFileUploader"] section {
    border: 1.5px dashed #D1D5DB !important;
    border-radius: 8px !important;
    background: #FFFFFF !important;
}

/* ── Dataframe / table ── */
.stDataFrame {
    border: 1px solid #E5E7EB !important;
    border-radius: 8px !important;
    overflow: hidden;
}

/* ── Dividers ── */
hr {
    border-color: #E5E7EB !important;
    margin: 1rem 0 !important;
}

/* ── Checkbox ── */
.stCheckbox label {
    font-size: 0.875rem !important;
    color: #374151 !important;
}

/* ── Number input ── */
.stNumberInput input {
    border-radius: 6px !important;
    border-color: #D1D5DB !important;
    font-size: 0.875rem !important;
}

/* ── Expander ── */
.streamlit-expanderHeader {
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    color: #374151 !important;
}

/* ── Suppress Streamlit rainbow top bar ── */
header[data-testid="stHeader"] {
    background: rgba(247, 248, 250, 0.95) !important;
    border-bottom: 1px solid #E5E7EB !important;
    backdrop-filter: blur(4px);
}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def well_ids():
    """All 120 well ids in row-major order."""
    return [f"{COL_LABELS[c]}{r + 1}" for r in range(ROWS) for c in range(COLS)]


def file_signature(uploaded_file) -> str:
    """Unique identity for an uploaded file based on its actual content."""
    data = uploaded_file.getvalue()
    chunk = 256 * 1024  # 256 KB
    if len(data) <= chunk * 2:
        sample = data
    else:
        sample = data[:chunk] + data[-chunk:]
    raw = uploaded_file.name.encode() + sample
    return hashlib.sha256(raw).hexdigest()[:16]


# ---------------------------------------------------------------------------
# STATE PERSISTENCE
# ---------------------------------------------------------------------------

def state_path(sig: str) -> str:
    return os.path.join(STATE_DIR, f"{sig}.json")


def load_state(sig: str) -> dict:
    path = state_path(sig)
    data: dict = {}
    if os.path.exists(path):
        try:
            with open(path) as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            data = {}
    if "wells" not in data or set(data["wells"].keys()) != set(well_ids()):
        data["wells"] = {w: "present" for w in well_ids()}
    data["calibration"] = normalize_calibration(data.get("calibration"))
    return data


def save_state(sig: str, data: dict) -> bool:
    try:
        with open(state_path(sig), "w") as f:
            json.dump(data, f, indent=2)
        return True
    except OSError:
        st.warning(
            "Could not write autosave file (read-only filesystem?). "
            "Your changes are kept for this session."
        )
        return False


# ---------------------------------------------------------------------------
# IMAGE DECODE — cached, lazy
# ---------------------------------------------------------------------------

@st.cache_data(max_entries=MAX_IMAGES, show_spinner=False)
def _decode_and_resize(_file_bytes: bytes, sig: str, max_w: int) -> bytes:
    img = Image.open(BytesIO(_file_bytes))
    if img.width > max_w:
        ratio = max_w / img.width
        img = img.resize((max_w, int(img.height * ratio)), Image.LANCZOS)
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def get_working_image(uploaded_file, sig: str) -> Image.Image:
    raw_bytes = uploaded_file.getvalue()
    png_bytes  = _decode_and_resize(raw_bytes, sig, MAX_DISPLAY_WIDTH)
    return Image.open(BytesIO(png_bytes))


# ---------------------------------------------------------------------------
# GRID MATH
# ---------------------------------------------------------------------------

def default_calibration(img_w: int, img_h: int):
    mx = img_w  * 0.04
    my = img_h  * 0.06
    return [[mx, my], [img_w - mx, img_h - my]]


def normalize_calibration(calibration) -> dict | None:
    if not calibration:
        return None
    if isinstance(calibration, list):
        return {"gap_after_row": 0, "points": calibration}
    if isinstance(calibration, dict):
        return {
            "gap_after_row": int(calibration.get("gap_after_row", 0) or 0),
            "points": calibration.get("points", []),
        }
    return None


def required_calibration_points(gap_after_row: int) -> int:
    return 4 if gap_after_row > 0 else 2


def calibration_instructions(gap_after_row: int) -> list[str]:
    if gap_after_row <= 0:
        return [
            "Click the **outer top-left corner** of the grid (just outside A1).",
            f"Click the **outer bottom-right corner** of the grid "
            f"(just outside {COL_LABELS[-1]}{ROWS}).",
        ]
    nxt = gap_after_row + 1
    return [
        "Click the **outer top-left corner** of the grid (just outside A1).",
        f"Click just **below row {gap_after_row}** — the bottom edge of the "
        f"last row before the gap (e.g. just below A{gap_after_row}).",
        f"Click just **above row {nxt}** — the top edge of the first row "
        f"after the gap (e.g. just above A{nxt}).",
        f"Click the **outer bottom-right corner** of the grid "
        f"(just outside {COL_LABELS[-1]}{ROWS}).",
    ]


def compute_cell_bounds(calibration, img_w: int, img_h: int):
    calib = normalize_calibration(calibration)
    gap_after_row = calib["gap_after_row"] if calib else 0
    pts = calib["points"] if calib else []
    req = required_calibration_points(gap_after_row)

    if len(pts) < req:
        pts = default_calibration(img_w, img_h)
        gap_after_row = 0

    if gap_after_row <= 0:
        (x1, y1), (x2, y2) = pts[0], pts[1]
        x1, x2 = min(x1, x2), max(x1, x2)
        y1, y2 = min(y1, y2), max(y1, y2)
        col_w = (x2 - x1) / COLS
        row_h = (y2 - y1) / ROWS
        bounds = {}
        for r in range(ROWS):
            for c in range(COLS):
                wid  = f"{COL_LABELS[c]}{r + 1}"
                cx0  = x1 + c * col_w
                cy0  = y1 + r * row_h
                bounds[wid] = (cx0, cy0, cx0 + col_w, cy0 + row_h)
        geometry = {
            "x1": x1, "x2": x2, "col_w": col_w,
            "bands": [
                {"y1": y1, "y2": y2, "row_h": row_h, "row_start": 0, "row_count": ROWS}
            ],
        }
        return bounds, geometry

    p1, p2, p3, p4 = pts
    x1, x2 = min(p1[0], p4[0]), max(p1[0], p4[0])
    y1a, y2a = sorted([p1[1], p2[1]])
    y1b, y2b = sorted([p3[1], p4[1]])

    row_count1 = gap_after_row
    row_count2 = ROWS - gap_after_row
    row_h1 = (y2a - y1a) / row_count1 if row_count1 > 0 else 0
    row_h2 = (y2b - y1b) / row_count2 if row_count2 > 0 else 0
    col_w  = (x2 - x1) / COLS

    bounds = {}
    for r in range(ROWS):
        if r < row_count1:
            y0  = y1a + r * row_h1
            y1c = y0 + row_h1
        else:
            r2  = r - row_count1
            y0  = y1b + r2 * row_h2
            y1c = y0 + row_h2
        for c in range(COLS):
            wid = f"{COL_LABELS[c]}{r + 1}"
            x0  = x1 + c * col_w
            bounds[wid] = (x0, y0, x0 + col_w, y1c)

    geometry = {
        "x1": x1, "x2": x2, "col_w": col_w,
        "bands": [
            {"y1": y1a, "y2": y2a, "row_h": row_h1, "row_start": 0, "row_count": row_count1},
            {"y1": y1b, "y2": y2b, "row_h": row_h2, "row_start": row_count1, "row_count": row_count2},
        ],
    }
    return bounds, geometry


def find_cell(x: float, y: float, geometry: dict) -> str | None:
    x1, x2, col_w = geometry["x1"], geometry["x2"], geometry["col_w"]
    if x < x1 or x > x2:
        return None
    col = min(int((x - x1) // col_w), COLS - 1) if col_w > 0 else 0
    for band in geometry["bands"]:
        if band["y1"] <= y <= band["y2"]:
            row_h = band["row_h"]
            local_row = min(int((y - band["y1"]) // row_h), band["row_count"] - 1) if row_h > 0 else 0
            row = band["row_start"] + local_row
            return f"{COL_LABELS[col]}{row + 1}"
    return None


# ---------------------------------------------------------------------------
# OVERLAY RENDERING
# ---------------------------------------------------------------------------

def _dashed_hline(draw, y, width, color, dash=10, gap=6):
    x = 0
    while x < width:
        x_end = min(x + dash, width)
        draw.line([(x, y), (x_end, y)], fill=color, width=2)
        x += dash + gap


def draw_grid_overlay(
    base_img: Image.Image,
    bounds: dict,
    wells: dict,
    show_labels: bool,
    calib_points=None,
    gap_after_row: int = 0,
    gap_band=None,
) -> Image.Image:
    img   = base_img.convert("RGBA")
    layer = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw  = ImageDraw.Draw(layer)
    font  = ImageFont.load_default()

    if gap_band is not None:
        y_top, y_bot = sorted(gap_band)
        draw.rectangle([0, y_top, img.width, y_bot], fill=(120, 120, 120, 55))

    for wid, (x0, y0, x1, y1) in bounds.items():
        present = wells.get(wid, "present") == "present"
        fill    = (*PRESENT_FILL, 80) if present else (*EMPTY_FILL, 130)
        draw.rectangle([x0, y0, x1, y1], fill=fill, outline=GRID_LINE)
        if show_labels:
            cx, cy = (x0 + x1) / 2, (y0 + y1) / 2
            draw.text((cx, cy), wid, fill=(0, 0, 0, 255), font=font, anchor="mm")

    if calib_points:
        if gap_after_row > 0 and len(calib_points) >= 3:
            y_top, y_bot = sorted([calib_points[1][1], calib_points[2][1]])
            draw.rectangle([0, y_top, img.width, y_bot], fill=(120, 120, 120, 70))

        for i, (x, y) in enumerate(calib_points, start=1):
            r = 9
            _dashed_hline(draw, y, img.width, (*CALIB_COLOR, 140))
            draw.ellipse(
                [x - r, y - r, x + r, y + r],
                outline=(*CALIB_COLOR, 255),
                width=3,
            )
            draw.text((x, y - r - 12), str(i), fill=(*CALIB_COLOR, 255), font=font, anchor="mm")

    return Image.alpha_composite(img, layer).convert("RGB")


# ---------------------------------------------------------------------------
# FREQUENCY / EXPORT DATA
# ---------------------------------------------------------------------------

def compute_frequency(sigs_and_names: list[tuple[str, str]]) -> dict:
    freq = {w: 0 for w in well_ids()}
    n_images = len(sigs_and_names)
    for sig, _ in sigs_and_names:
        data = load_state(sig)
        for wid, status in data["wells"].items():
            if status == "empty":
                freq[wid] += 1
    return freq, n_images


_GRADIENT_CMAP = mcolors.LinearSegmentedColormap.from_list(
    "green_yellow_orange_red",
    ["#63BE7B", "#FFEB84", "#FFA500", "#F8696B"],
)


def gradient_hex(count: int, maximum: int) -> str:
    ratio = (count / maximum) if maximum > 0 else 0.0
    ratio = max(0.0, min(1.0, ratio))
    r, g, b, _a = _GRADIENT_CMAP(ratio)
    return f"{int(round(r * 255)):02X}{int(round(g * 255)):02X}{int(round(b * 255)):02X}"


def readable_text_hex(bg_hex: str) -> str:
    r, g, b = int(bg_hex[0:2], 16), int(bg_hex[2:4], 16), int(bg_hex[4:6], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return "000000" if luminance > 0.6 else "FFFFFF"


def build_heatmap_figure(freq: dict, n_images: int) -> plt.Figure:
    grid = np.zeros((ROWS, COLS), dtype=float)
    for r in range(ROWS):
        for c in range(COLS):
            wid = f"{COL_LABELS[c]}{r + 1}"
            grid[r, c] = freq.get(wid, 0)

    fig, ax = plt.subplots(figsize=(14, 6))
    cmap = _GRADIENT_CMAP
    im = ax.imshow(grid, cmap=cmap, aspect="auto",
                   vmin=0, vmax=max(n_images, 1))

    for r in range(ROWS):
        for c in range(COLS):
            val = int(grid[r, c])
            pct = (val / n_images * 100) if n_images > 0 else 0
            text_hex = readable_text_hex(gradient_hex(val, max(n_images, 1)))
            color = "white" if text_hex == "FFFFFF" else "black"
            ax.text(c, r, f"{val}\n({pct:.0f}%)", ha="center", va="center",
                    fontsize=7, color=color, fontweight="bold")

    ax.set_xticks(range(COLS))
    ax.set_xticklabels(list(COL_LABELS), fontsize=9)
    ax.set_yticks(range(ROWS))
    ax.set_yticklabels([str(i + 1) for i in range(ROWS)], fontsize=9)
    ax.set_xlabel("Column", fontsize=11)
    ax.set_ylabel("Row", fontsize=11)
    ax.set_title(
        f"Missing Coordinate Frequency Heatmap\n"
        f"{n_images} mold{'s' if n_images != 1 else ''} analyzed",
        fontsize=13, fontweight="bold"
    )

    cbar = fig.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label("# Times Missing", fontsize=9)
    plt.tight_layout()
    return fig


def build_excel_export(freq: dict, n_images: int) -> bytes:
    wb = openpyxl.Workbook()

    # ---- SHEET 1: Heatmap ------------------------------------------------
    ws_heat = wb.active
    ws_heat.title = "Heatmap"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    label_font   = Font(name="Arial", bold=True, size=10)
    cell_font    = Font(name="Arial", size=9)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_heat.merge_cells("A1:Q1")
    title_cell = ws_heat["A1"]
    title_cell.value = f"Missing Coordinate Frequency Heatmap — {n_images} Mold(s) Analyzed"
    title_cell.font  = Font(name="Arial", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="2C3E50")
    title_cell.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws_heat.row_dimensions[1].height = 30

    ws_heat["A2"].value = "Row \\ Col"
    ws_heat["A2"].font  = header_font
    ws_heat["A2"].fill  = PatternFill("solid", fgColor="34495E")
    ws_heat["A2"].alignment = center
    ws_heat["A2"].border = border

    for c_idx, col_letter in enumerate(COL_LABELS):
        cell = ws_heat.cell(row=2, column=c_idx + 2)
        cell.value = col_letter
        cell.font  = header_font
        cell.fill  = PatternFill("solid", fgColor="34495E")
        cell.alignment = center
        cell.border = border

    ws_heat.row_dimensions[2].height = 22

    max_freq = max(freq.values()) if freq else 1

    for r_idx in range(ROWS):
        row_num = r_idx + 3
        ws_heat.row_dimensions[row_num].height = 28

        label_cell = ws_heat.cell(row=row_num, column=1)
        label_cell.value = str(r_idx + 1)
        label_cell.font  = label_font
        label_cell.fill  = PatternFill("solid", fgColor="ECF0F1")
        label_cell.alignment = center
        label_cell.border = border

        for c_idx in range(COLS):
            wid   = f"{COL_LABELS[c_idx]}{r_idx + 1}"
            count = freq.get(wid, 0)
            pct   = (count / n_images * 100) if n_images > 0 else 0
            hex_color = gradient_hex(count, max_freq)

            data_cell = ws_heat.cell(row=row_num, column=c_idx + 2)
            data_cell.value = f"{count} ({pct:.0f}%)"
            data_cell.font  = Font(
                name="Arial", size=8, bold=count > 0,
                color=readable_text_hex(hex_color)
            )
            data_cell.fill  = PatternFill("solid", fgColor=hex_color)
            data_cell.alignment = center
            data_cell.border = border

    ws_heat.column_dimensions["A"].width = 10
    for c_idx in range(COLS):
        ws_heat.column_dimensions[get_column_letter(c_idx + 2)].width = 10

    note_row = ROWS + 4
    ws_heat.merge_cells(f"A{note_row}:Q{note_row}")
    note = ws_heat.cell(row=note_row, column=1)
    note.value = (
        f"Each cell shows: count of molds where this coordinate was missing "
        f"(percentage of {n_images} total molds). "
        f"Color follows a green → yellow → orange → red scale: "
        f"green = rarely missing, red = frequently missing."
    )
    note.font = Font(name="Arial", italic=True, size=9, color="555555")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_heat.row_dimensions[note_row].height = 30

    # ---- SHEET 2: Frequency Table ----------------------------------------
    ws_table = wb.create_sheet("Frequency Table")

    ws_table.merge_cells("A1:E1")
    t = ws_table["A1"]
    t.value = f"Coordinate Flag Frequency Table — {n_images} Mold(s) Analyzed"
    t.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor="2C3E50")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_table.row_dimensions[1].height = 30

    headers = ["Coordinate", "Column", "Row", "Times Missing", "% Missing"]
    header_fills = ["34495E"] * 5
    for col_i, (h, fill) in enumerate(zip(headers, header_fills), start=1):
        cell = ws_table.cell(row=2, column=col_i)
        cell.value = h
        cell.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor=fill)
        cell.alignment = center
        cell.border = border
    ws_table.row_dimensions[2].height = 22

    all_wids = well_ids()
    sorted_wids = sorted(all_wids, key=lambda w: -freq.get(w, 0))

    for row_i, wid in enumerate(sorted_wids, start=3):
        count = freq.get(wid, 0)
        pct   = (count / n_images * 100) if n_images > 0 else 0
        col_letter = wid[0]
        row_number = wid[1:]

        hex_color = gradient_hex(count, max_freq)
        text_color = readable_text_hex(hex_color)

        vals = [wid, col_letter, row_number, count, round(pct, 1)]
        for col_i, val in enumerate(vals, start=1):
            cell = ws_table.cell(row=row_i, column=col_i)
            cell.value = val
            cell.border = border
            cell.alignment = center

            if col_i == 4:
                cell.fill = PatternFill("solid", fgColor=hex_color)
                cell.font = Font(name="Arial", size=10, bold=True, color=text_color)
            elif col_i == 5:
                cell.fill = PatternFill("solid", fgColor=hex_color)
                cell.font = Font(name="Arial", size=10, color=text_color)
                cell.number_format = "0.0%"
                cell.value = pct / 100
            else:
                cell.font = Font(name="Arial", size=10)
                if row_i % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F8F9FA")

        ws_table.row_dimensions[row_i].height = 18

    col_widths = [14, 10, 8, 16, 14]
    for col_i, w in enumerate(col_widths, start=1):
        ws_table.column_dimensions[get_column_letter(col_i)].width = w

    footer_row = len(sorted_wids) + 4
    ws_table.merge_cells(f"A{footer_row}:E{footer_row}")
    footer = ws_table.cell(row=footer_row, column=1)
    total_missing = sum(freq.values())
    footer.value = (
        f"Total missing observations across all coordinates: {total_missing}  |  "
        f"Molds analyzed: {n_images}  |  Sorted by frequency (highest first)"
    )
    footer.font = Font(name="Arial", italic=True, size=9, color="555555")
    footer.alignment = Alignment(horizontal="left", vertical="center")
    ws_table.row_dimensions[footer_row].height = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# UI
# ===========================================================================

st.title("Mold Coordinate Tracker")
st.caption(
    f"Upload up to {MAX_IMAGES} mold photos. Calibrate the grid on the first image, "
    "mark coordinates on each subsequent image, then export a frequency heatmap."
)

# ---------------------------------------------------------------------------
# File upload
# ---------------------------------------------------------------------------
uploaded_files = st.file_uploader(
    f"Upload mold images (max {MAX_IMAGES})",
    type=["png", "jpg", "jpeg"],
    accept_multiple_files=True,
)

if not uploaded_files:
    st.info("Upload one or more images to get started.")
    st.stop()

if len(uploaded_files) > MAX_IMAGES:
    st.warning(
        f"You uploaded {len(uploaded_files)} images but the limit is {MAX_IMAGES}. "
        f"Only the first {MAX_IMAGES} will be used.",
    )
    uploaded_files = uploaded_files[:MAX_IMAGES]

_raw_sigs = [file_signature(f) for f in uploaded_files]
sigs_and_names: list[tuple[str, str]] = []
_seen: dict[str, int] = {}
for i, (raw_sig, f) in enumerate(zip(_raw_sigs, uploaded_files)):
    if raw_sig in _seen:
        unique_sig = f"{raw_sig}_{i}"
    else:
        unique_sig = raw_sig
    _seen[raw_sig] = i
    sigs_and_names.append((unique_sig, f.name))

sig_to_file = {sig: f for (sig, _), f in zip(sigs_and_names, uploaded_files)}
all_states: dict[str, dict] = {sig: load_state(sig) for sig, _ in sigs_and_names}
sig_list  = [sig for sig, _ in sigs_and_names]
name_map  = {sig: name for sig, name in sigs_and_names}

n_total = len(sig_list)

# ---------------------------------------------------------------------------
# Shared calibration
# ---------------------------------------------------------------------------
first_sig   = sig_list[0]
first_calib = all_states[first_sig].get("calibration")
if first_calib:
    _req = required_calibration_points(first_calib.get("gap_after_row", 0))
    if len(first_calib.get("points", [])) == _req:
        for sig in sig_list[1:]:
            if not all_states[sig].get("calibration"):
                all_states[sig]["calibration"] = {
                    "gap_after_row": first_calib["gap_after_row"],
                    "points": list(first_calib["points"]),
                }
                save_state(sig, all_states[sig])

# ---------------------------------------------------------------------------
# Session state: current image index
# ---------------------------------------------------------------------------
if "current_idx" not in st.session_state:
    st.session_state.current_idx = 0
st.session_state.current_idx = max(0, min(st.session_state.current_idx, n_total - 1))

# ---------------------------------------------------------------------------
# Top navigation tabs
# ---------------------------------------------------------------------------
tab_annotate, tab_export = st.tabs(["Annotate Images", "Export Results"])

# ===========================================================================
# TAB 1 — ANNOTATE
# ===========================================================================
with tab_annotate:

    idx = st.session_state.current_idx
    active_sig   = sig_list[idx]
    active_file  = sig_to_file[active_sig]
    active_state = all_states[active_sig]

    work_img     = get_working_image(active_file, active_sig)
    img_w, img_h = work_img.size
    bounds, geometry = compute_cell_bounds(active_state["calibration"], img_w, img_h)

    # --- Header row -------------------------------------------------------
    h_col1, h_col2, h_col3 = st.columns([1, 4, 1])
    with h_col1:
        if st.button("Previous", key="btn_prev_top", disabled=(idx == 0), use_container_width=True):
            st.session_state.current_idx -= 1
            st.rerun()
    with h_col2:
        st.markdown(
            f"<div style='text-align:center;font-size:0.9rem;padding-top:6px;"
            f"color:#6B7280;font-family:Inter,sans-serif;'>"
            f"<span style='font-weight:600;color:#111827;'>Image {idx + 1} of {n_total}</span>"
            f"&nbsp;&nbsp;·&nbsp;&nbsp;"
            f"<code style='background:#F3F4F6;padding:2px 6px;border-radius:4px;"
            f"font-size:0.82rem;color:#374151;'>{name_map[active_sig]}</code>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with h_col3:
        if st.button("Next", key="btn_next_top", disabled=(idx == n_total - 1), use_container_width=True):
            st.session_state.current_idx += 1
            st.rerun()

    st.progress((idx + 1) / n_total)

    # --- Sidebar ----------------------------------------------------------
    with st.sidebar:
        if st.session_state.get("confirm_full_reset"):
            st.error(
                "This deletes ALL saved calibration and coordinate data for "
                "every image. This cannot be undone."
            )
            rc1, rc2 = st.columns(2)
            with rc1:
                if st.button("Confirm reset", key="btn_confirm_reset", use_container_width=True, type="primary"):
                    for fname in os.listdir(STATE_DIR):
                        try:
                            os.remove(os.path.join(STATE_DIR, fname))
                        except OSError:
                            pass
                    for k in list(st.session_state.keys()):
                        del st.session_state[k]
                    st.rerun()
            with rc2:
                if st.button("Cancel", key="btn_cancel_reset", use_container_width=True):
                    st.session_state["confirm_full_reset"] = False
                    st.rerun()
        else:
            if st.button("Reset — start from scratch", key="btn_start_reset", use_container_width=True):
                st.session_state["confirm_full_reset"] = True
                st.rerun()

        st.divider()
        st.header("Controls")
        show_labels = st.checkbox("Show coordinate labels", value=True)

        st.divider()
        present_count = sum(1 for v in active_state["wells"].values() if v == "present")
        empty_count   = len(active_state["wells"]) - present_count
        c1, c2 = st.columns(2)
        c1.metric("Present", present_count)
        c2.metric("Empty",   empty_count)

        st.divider()
        if st.button("Reset all coordinates to Present", key="btn_reset_wells", use_container_width=True):
            active_state["wells"] = {w: "present" for w in well_ids()}
            save_state(active_sig, active_state)
            st.rerun()
        if st.button("Reset calibration", key="btn_reset_calib", use_container_width=True):
            active_state["calibration"] = None
            save_state(active_sig, active_state)
            st.rerun()

        st.divider()
        st.download_button(
            "Download this image (CSV)",
            pd.DataFrame(
                [{"well": w, "status": s} for w, s in active_state["wells"].items()]
            ).to_csv(index=False).encode("utf-8"),
            file_name=f"mold_{active_sig}_wells.csv",
            mime="text/csv",
            use_container_width=True,
        )
        st.download_button(
            "Download this image (JSON)",
            json.dumps(active_state, indent=2).encode("utf-8"),
            file_name=f"mold_{active_sig}_state.json",
            mime="application/json",
            use_container_width=True,
        )

        restore_file = st.file_uploader(
            "Restore state (JSON)",
            type=["json"],
            key="restore",
        )
        if restore_file is not None:
            try:
                restored = json.load(restore_file)
                if "wells" in restored:
                    active_state["wells"].update(restored["wells"])
                if restored.get("calibration"):
                    active_state["calibration"] = normalize_calibration(restored["calibration"])
                save_state(active_sig, active_state)
                st.success("State restored.")
                st.rerun()
            except (json.JSONDecodeError, KeyError):
                st.error("Invalid state file.")

    # --- Mode: calibrate or mark ------------------------------------------
    active_calib = active_state["calibration"]
    needs_calibration = (
        active_calib is None
        or len(active_calib.get("points", [])) < required_calibration_points(
            active_calib.get("gap_after_row", 0)
        )
    )

    if needs_calibration:
        # ---- CALIBRATION MODE ----------------------------------------
        st.subheader("Step 1 — Calibrate the grid")

        if active_calib is None:
            st.info(
                "This image needs calibration. Set which row the gap falls after, "
                "then follow the on-screen prompts to place the corner points."
            )

            prior_gap = st.session_state.get("gap_after_row_choice") or (ROWS // 2)
            gap_after_row = st.number_input(
                f"Gap occurs after row # (1–{ROWS - 1})",
                min_value=1,
                max_value=ROWS - 1,
                value=prior_gap,
                step=1,
                key=f"gap_row_{active_sig}",
            )

            st.image(work_img, use_container_width=True, caption="Preview — not yet calibrated")

            if st.button("Start calibration", key=f"start_calib_{active_sig}", type="primary"):
                gap_after_row = int(gap_after_row)
                st.session_state["gap_after_row_choice"] = gap_after_row
                active_state["calibration"] = {"gap_after_row": gap_after_row, "points": []}
                save_state(active_sig, active_state)
                st.rerun()

        else:
            gap_after_row = active_calib.get("gap_after_row", 0)
            calib_pts = active_calib.get("points", [])
            req = required_calibration_points(gap_after_row)
            instructions = calibration_instructions(gap_after_row)

            st.info(f"Point {len(calib_pts) + 1} of {req}: {instructions[len(calib_pts)]}")
            if gap_after_row > 0:
                st.caption(
                    f"Layout: rows 1–{gap_after_row} in the upper block, "
                    f"rows {gap_after_row + 1}–{ROWS} in the lower block, "
                    "with a gap between them."
                )

            overlay = draw_grid_overlay(
                work_img, bounds, active_state["wells"], show_labels,
                calib_points=calib_pts, gap_after_row=gap_after_row,
            )
            click = streamlit_image_coordinates(overlay, key=f"calib_{active_sig}")

            last_key = f"last_calib_click_{active_sig}"
            if click is not None:
                sig_xy = (click.get("x"), click.get("y"))
                if sig_xy != (None, None) and st.session_state.get(last_key) != sig_xy:
                    st.session_state[last_key] = sig_xy
                    pts = list(calib_pts)
                    if len(pts) >= req:
                        pts = []
                    pts.append([sig_xy[0], sig_xy[1]])
                    active_state["calibration"]["points"] = pts
                    save_state(active_sig, active_state)
                    st.rerun()

            with st.expander("Restart calibration"):
                st.caption("Clears the points placed so far for this image (keeps the gap setting).")
                if st.button("Clear points and restart", key=f"restart_pts_{active_sig}"):
                    active_state["calibration"] = {"gap_after_row": gap_after_row, "points": []}
                    save_state(active_sig, active_state)
                    st.rerun()
                st.caption("Or change the gap layout entirely for this image:")
                if st.button("Change gap layout", key=f"change_gap_{active_sig}"):
                    active_state["calibration"] = None
                    save_state(active_sig, active_state)
                    st.rerun()

    else:
        # ---- MARK COORDINATES MODE -----------------------------------
        st.subheader("Step 2 — Click a cell to toggle Empty / Present")

        _active_gap = active_calib.get("gap_after_row", 0) if active_calib else 0
        gap_note = f"  ·  gap after row {_active_gap}" if _active_gap > 0 else ""
        st.caption(
            f"Green = present   Red = empty   ·   "
            f"Grid: {COLS} columns (A–{COL_LABELS[-1]}) × {ROWS} rows{gap_note}"
        )

        gap_band = None
        if len(geometry["bands"]) == 2:
            gap_band = (geometry["bands"][0]["y2"], geometry["bands"][1]["y1"])

        overlay = draw_grid_overlay(
            work_img, bounds, active_state["wells"], show_labels, gap_band=gap_band
        )
        click = streamlit_image_coordinates(overlay, key=f"mark_{active_sig}")

        last_key = f"last_mark_click_{active_sig}"
        if click is not None:
            sig_xy = (click.get("x"), click.get("y"))
            if sig_xy != (None, None) and st.session_state.get(last_key) != sig_xy:
                st.session_state[last_key] = sig_xy
                wid = find_cell(sig_xy[0], sig_xy[1], geometry)
                if wid:
                    current = active_state["wells"][wid]
                    active_state["wells"][wid] = "empty" if current == "present" else "present"
                    save_state(active_sig, active_state)
                    st.rerun()
                else:
                    st.toast("Click landed outside the calibrated grid.")

        with st.expander("Show empty coordinate list"):
            empty_wells = [w for w, s in active_state["wells"].items() if s == "empty"]
            st.write(", ".join(empty_wells) if empty_wells else "No coordinates marked empty yet.")

        # Bottom navigation
        st.divider()
        nav1, nav2, nav3 = st.columns([1, 6, 1])
        with nav1:
            if st.button("Previous", key="btn_prev_bottom", disabled=(idx == 0), use_container_width=True):
                st.session_state.current_idx -= 1
                st.rerun()
        with nav2:
            progress_labels = []
            for i, (s, n) in enumerate(sigs_and_names):
                state_i = all_states[s]
                _c = state_i.get("calibration")
                calib_ok = bool(_c) and len(_c.get("points", [])) >= required_calibration_points(_c.get("gap_after_row", 0))
                status_marker = "[done]" if calib_ok else "[pending]"
                progress_labels.append(f"{status_marker} {i + 1}")
            st.caption("  ".join(progress_labels))
        with nav3:
            if st.button("Next", key="btn_next_bottom", disabled=(idx == n_total - 1), use_container_width=True):
                st.session_state.current_idx += 1
                st.rerun()

            if idx == n_total - 1:
                st.success("All images reviewed. Open the Export Results tab to download your report.")


# ===========================================================================
# TAB 2 — EXPORT
# ===========================================================================
with tab_export:

    freq, n_images = compute_frequency(sigs_and_names)

    def _is_calibrated(sig):
        c = all_states[sig].get("calibration")
        if not c:
            return False
        return len(c.get("points", [])) >= required_calibration_points(c.get("gap_after_row", 0))

    n_calibrated = sum(1 for sig, _ in sigs_and_names if _is_calibrated(sig))

    st.subheader(f"Results — {n_images} mold(s) uploaded, {n_calibrated} calibrated")

    if n_images == 0:
        st.info("No images uploaded yet.")
    else:
        # ---- Heatmap preview -------------------------------------------
        st.markdown("### Missing Coordinate Frequency Heatmap")
        st.caption(
            "Color intensity indicates how often each coordinate was marked empty "
            "across all analyzed molds."
        )
        fig = build_heatmap_figure(freq, n_images)
        st.pyplot(fig, use_container_width=True)
        plt.close(fig)

        heatmap_buf = BytesIO()
        fig2 = build_heatmap_figure(freq, n_images)
        fig2.savefig(heatmap_buf, format="PNG", dpi=150, bbox_inches="tight")
        plt.close(fig2)
        heatmap_buf.seek(0)

        st.download_button(
            "Download Heatmap (PNG)",
            heatmap_buf.getvalue(),
            file_name="mold_heatmap.png",
            mime="image/png",
            use_container_width=False,
        )

        st.divider()

        # ---- Frequency table -------------------------------------------
        st.markdown("### Coordinate Flag Frequency Table")
        rows = []
        for wid in well_ids():
            count = freq.get(wid, 0)
            pct   = (count / n_images * 100) if n_images > 0 else 0
            rows.append({
                "Coordinate": wid,
                "Column": wid[0],
                "Row": wid[1:],
                "Times Missing": count,
                "% Missing": round(pct, 1),
            })

        df_freq = pd.DataFrame(rows).sort_values("Times Missing", ascending=False)
        st.dataframe(
            df_freq,
            use_container_width=True,
            hide_index=True,
            column_config={
                "% Missing": st.column_config.NumberColumn(format="%.1f%%"),
                "Times Missing": st.column_config.ProgressColumn(
                    min_value=0, max_value=n_images, format="%d"
                ),
            },
        )

        st.divider()

        # ---- Excel export ----------------------------------------------
        st.markdown("### Download Excel Report")
        st.caption(
            "The workbook contains two sheets: Heatmap (color-coded grid) "
            "and Frequency Table (all 120 coordinates sorted by missing frequency)."
        )

        excel_bytes = build_excel_export(freq, n_images)
        st.download_button(
            "Download Excel Report (.xlsx)",
            excel_bytes,
            file_name="mold_coordinate_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )
